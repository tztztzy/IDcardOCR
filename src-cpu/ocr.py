# -*- coding: utf-8 -*-
import os
import shutil
from pathlib import Path
import re
import glob
from PIL import Image, ImageEnhance
from paddleocr import PaddleOCR, draw_ocr
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import Workbook

# 初始化 Excel 工作簿
excel_file = "recognized_id_info.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "性别", "民族", "出生日期", "住址", "身份证号"])
    wb.save(excel_file)

class IDRecognitionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("身份证信息识别系统 Designed by TZY")
        self.root.geometry("800x600")
        
        # 创建主界面布局
        self.create_widgets()
        
        # 初始化处理状态
        self.processing = False

    def create_widgets(self):
        # 顶部按钮区域
        control_frame = ttk.Frame(self.root)
        control_frame.pack(pady=20, fill=tk.X)
        
        # 单个处理按钮
        self.single_btn = ttk.Button(
            control_frame,
            text="选择单个文件",
            command=self.recognize_single,
            width=15
        )
        self.single_btn.pack(side=tk.LEFT, padx=10)
        
        # 批量处理按钮
        self.batch_btn = ttk.Button(
            control_frame,
            text="批量处理文件夹",
            command=self.recognize_batch,
            width=15
        )
        self.batch_btn.pack(side=tk.LEFT, padx=10)
        
        # 进度条
        self.progress = ttk.Progressbar(
            self.root,
            orient=tk.HORIZONTAL,
            mode='determinate'
        )
        self.progress.pack(fill=tk.X, padx=20, pady=10)
        
        # 版本和性能提示
        info_label = ttk.Label(
            self.root,
            text="身份证信息识别系统 By TZY v0.0.1\n--------------------------------------------------------------------------------------------\n由于运算量巨大，程序运行过程中会大量占用CPU，电脑会较卡顿，程序可能未响应，此为正常现象，运行完毕后即恢复正常\n二值化处理图像和标注图像存储于/tmp目录中，同名文件只会创建一次。识别结果保存到recognized_id_info.xlsx文件中\n为保证识别准确性，使用模型为全尺寸模型，若电脑配置过低可以更换小尺寸蒸馏模型，但精确度会降低。\n识别过程中不要打开excel文件，否则会无法写入!!!!\n--------------------------------------------------------------------------------------------",
            wraplength=700,
            justify=tk.CENTER
        )
        info_label.pack(pady=(0, 5))
        
        # 日志显示区域
        log_frame = ttk.LabelFrame(self.root, text="处理日志")
        log_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(
            log_frame,
            height=8,
            state=tk.DISABLED,
            wrap=tk.WORD
        )
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 状态栏
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(
            self.root,
            textvariable=self.status_var,
            relief=tk.SUNKEN
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.update_status("就绪")

    def update_status(self, message):
        self.status_var.set(message)
        self.root.update_idletasks()

    def log_message(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def toggle_buttons(self, state):
        """切换按钮状态"""
        state = tk.NORMAL if state else tk.DISABLED
        self.single_btn.config(state=state)
        self.batch_btn.config(state=state)

    def recognize_single(self):
        if self.processing: return
        file_path = filedialog.askopenfilename(
            title="选择身份证图片",
            filetypes=[("图片文件", "*.jpg;*.png;*.jpeg;*.bmp")]
        )
        if file_path:
            self.process_files([file_path])

    def recognize_batch(self):
        if self.processing: return
        folder_path = filedialog.askdirectory(title="选择图片文件夹")
        if folder_path:
            image_files = self.get_image_files(folder_path)
            if image_files:
                self.process_files(image_files)
            else:
                messagebox.showwarning("警告", "未找到支持的图片文件")

    def get_image_files(self, folder):
        extensions = ("*.jpg", "*.png", "*.jpeg", "*.bmp")
        return [f for ext in extensions for f in glob.glob(os.path.join(folder, ext))]

    def process_files(self, files):
        self.processing = True
        self.toggle_buttons(False)
        self.progress["maximum"] = len(files)
        
        import threading
        
        def process_thread():
            try:
                for i, file_path in enumerate(files, 1):
                    self.root.after(0, lambda: self.update_status(f"正在处理: {os.path.basename(file_path)} ({i}/{len(files)})"))
                    self.root.after(0, lambda: self.progress.configure(value=i))
                    self.root.after(0, lambda: self.log_message(f"▶ 开始处理: {file_path}"))
                    
                    result = self.recognize_image(file_path)
                    self.append_to_excel(result)
                    
                    log_entry = "  ✔ 识别结果: " + " | ".join(
                        [f"{k}:{v if v else '未识别'}" for k, v in result.items()]
                    )
                    self.root.after(0, lambda: self.log_message(log_entry))
                    
                self.root.after(0, lambda: messagebox.showinfo("完成", f"成功处理 {len(files)} 个文件"))
            except Exception as e:
                self.root.after(0, lambda: self.log_message(f"  ✖ 处理失败: {str(e)}"))
                self.root.after(0, lambda: messagebox.showerror("错误", f"处理过程中发生错误: {str(e)}"))
            finally:
                self.processing = False
                self.root.after(0, lambda: self.toggle_buttons(True))
                self.root.after(0, lambda: self.progress.configure(value=0))
                self.root.after(0, lambda: self.update_status("就绪"))
        
        threading.Thread(target=process_thread, daemon=True).start()

    def recognize_image(self, image_path):
        """图像识别核心逻辑"""
        try:
            # 图像预处理
            img = Image.open(image_path).convert('L')
            img = ImageEnhance.Contrast(img).enhance(2.0)
            img = ImageEnhance.Sharpness(img).enhance(2.0)
            enhanced_path = f"tmp/enhanced_{os.path.basename(image_path)}"
            img.save(enhanced_path)

            # OCR识别
            ocr = PaddleOCR(lang='ch')
            result = ocr.ocr(enhanced_path, cls=False)
            full_text = ''.join([item[1][0] for item in result[0]])
            
            # 信息提取
            patterns = {
                "姓名": [r'姓名(.+?)(:?性|别)', r'名(.+?)(:?性|别)', r'姓(.+?)(:?性|别)'],
                "性别": [r'性别([男女])', r'别([男女])', r'性([男女])'],
                "民族": [r'民族(.+?)(:?出|生)', r'族(.+?)(:?出|生)',r'民(.+?)(:?出|生)'],
                "出生日期": [r'出生(.+?)(:?住|址)', r'生(.+?)(:?住|址)',r'出(.+?)(:?住|址)'],
                "住址": [r'住址(.+?)(:?公|民)', r'址(.+?)(:?公|民)',r'住(.+?)(:?公|民)'],
                "身份证号": r'\d{17}[\dXx]'
            }
            
            result_data = {}
            for field, pattern in patterns.items():
                if field == "身份证号":
                    matches = re.findall(pattern, full_text)
                    result_data[field] = matches[0] if matches else None
                else:
                    for p in (pattern if isinstance(pattern, list) else [pattern]):
                        match = re.search(p, full_text)
                        if match:
                            result_data[field] = match.group(1)
                            break
                    else:
                        result_data[field] = None
                if field == "姓名" and result_data[field] and len(result_data[field]) == 1:
                    name_match = re.search(r'([^姓]+)(?:姓|名)', full_text)
                    if name_match:
                        result_data[field] = name_match.group(1).strip()
                if field == "住址" and result_data[field]:
                    result_data[field] = re.sub(r'\d{17}[\dXx]', '', result_data[field])
            
            # 保存标注图像
            self.save_annotated_image(enhanced_path, result)
            return result_data
            
        except Exception as e:
            raise RuntimeError(f"图像处理失败: {str(e)}")

    def save_annotated_image(self, img_path, ocr_result):
        """保存标注后的图像"""
        image = Image.open(img_path).convert('RGB')
        boxes = [line[0] for line in ocr_result[0]]
        txts = [line[1][0] for line in ocr_result[0]]
        scores = [line[1][1] for line in ocr_result[0]]
        
        annotated = draw_ocr(image, boxes, txts, scores)
        output_path = os.path.join("tmp", f"annotated_{os.path.basename(img_path)}")  
        Image.fromarray(annotated).save(output_path)

    def append_to_excel(self, data):
        """数据追加到Excel"""
        try:
            # 检查数据有效性
            if not data or not data.get("身份证号"):
                self.log_message("  ✖ 数据无效，跳过写入")
                return
                
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # 检查是否已存在相同身份证号的记录
            id_col = 6  # 身份证号在第6列
            for row in ws.iter_rows(values_only=True):
                if row and row[id_col-1] == data["身份证号"]:
                    self.log_message(f"  ⚠ 已存在相同身份证号记录: {data['身份证号']}")
                    return
            
            # 写入新记录
            ws.append([
                data["姓名"],
                data["性别"],
                data["民族"],
                data["出生日期"],
                data["住址"],
                data["身份证号"]
            ])
            wb.save(excel_file)
            self.log_message("  ✔ 数据已成功写入Excel")
        except Exception as e:
            self.log_message(f"  ✖ 保存到Excel失败: {str(e)}")
            raise RuntimeError(f"保存到Excel失败: {str(e)}")

if __name__ == "__main__":
    try:
        username = os.environ.get("USERNAME") or os.environ.get("USER")
        if not username:
            raise ValueError("无法获取用户名")
        src_dir = Path(__file__).parent / "paddleocr"
        dest_dir = Path(f"C:/Users/{username}/.paddleocr")

        dest_dir.mkdir(parents=True, exist_ok=True)

        if hasattr(shutil, "copytree"):
            shutil.copytree(src_dir, dest_dir / "paddlemodel", dirs_exist_ok=True)
        else:
            if (dest_dir / "paddlemodel").exists():
                shutil.rmtree(dest_dir / "paddlemodel")
            shutil.copytree(src_dir, dest_dir / "paddlemodel")
    except Exception as e:
        messagebox.showerror("初始化错误", f"模型复制失败: {str(e)}")
        exit()

    root = tk.Tk()
    app = IDRecognitionApp(root)
    root.mainloop()